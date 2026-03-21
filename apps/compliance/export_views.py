import hashlib
import json
from datetime import datetime, timedelta
from django.utils import timezone
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import ImmutableAuditLog
from .audit_chain import verify_audit_chain

def generate_audit_report_pdf(logs, workspace_id, chain_status):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Header
    title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=16, spaceAfter=6)
    story.append(Paragraph("LAPORAN AUDIT KOMUNIKASI", title_style))
    story.append(Paragraph(f"BlackMess Enterprise Platform", styles['Heading2']))
    story.append(Spacer(1, 0.3*cm))

    # Info laporan
    info_data = [
        ['Workspace ID', workspace_id],
        ['Tanggal Laporan', timezone.now().strftime('%d/%m/%Y %H:%M:%S WIB')],
        ['Total Log', str(len(logs))],
        ['Integritas Chain', 'VALID' if chain_status['valid'] else 'RUSAK - KEMUNGKINAN MANIPULASI!'],
        ['Standar', 'OJK POJK No.11/POJK.03/2022 | BI PBI No.23/6/PBI/2021'],
    ]

    info_table = Table(info_data, colWidths=[5*cm, 11*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.grey),
        ('TEXTCOLOR', (0,0), (0,-1), colors.white),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5*cm))

    # Tabel log
    story.append(Paragraph("RIWAYAT AKTIVITAS KOMUNIKASI", styles['Heading2']))
    story.append(Spacer(1, 0.2*cm))

    headers = ['No', 'Pengirim', 'Penerima', 'Aksi', 'Channel', 'IP', 'Waktu', 'Hash']
    table_data = [headers]

    for i, log in enumerate(logs[:500], 1):
        table_data.append([
            str(i),
            log.get('sender_id', '')[:15],
            log.get('receiver_id', '')[:15],
            log.get('action', ''),
            log.get('channel', '')[:20],
            log.get('ip_address', '') or '-',
            str(log.get('created_at', ''))[:16],
            log.get('chain_hash', '')[:12] + '...',
        ])

    log_table = Table(table_data, colWidths=[0.8*cm, 2.5*cm, 2.5*cm, 1.5*cm, 2.5*cm, 2.5*cm, 2.8*cm, 2.4*cm])
    log_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 7),
        ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
        ('PADDING', (0,0), (-1,-1), 4),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.lightgrey]),
    ]))
    story.append(log_table)

    # Footer
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        f"Dokumen ini digenerate otomatis oleh BlackMess Compliance System pada "
        f"{timezone.now().strftime('%d/%m/%Y %H:%M:%S WIB')}. "
        f"Hash dokumen: {hashlib.sha256(str(table_data).encode()).hexdigest()[:32]}",
        styles['Normal']
    ))

    doc.build(story)
    buffer.seek(0)
    return buffer

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_audit_pdf(request):
    """Export audit log ke PDF untuk OJK/BI"""
    workspace_id = request.query_params.get('workspace_id', 'default')
    days = int(request.query_params.get('days', 30))
    
    since = timezone.now() - timedelta(days=days)
    logs = list(ImmutableAuditLog.objects.filter(
        workspace_id=workspace_id,
        created_at__gte=since
    ).order_by('-created_at').values(
        'sender_id', 'receiver_id', 'action', 'channel',
        'ip_address', 'created_at', 'chain_hash', 'message_hash'
    ))
    
    chain_status = verify_audit_chain(workspace_id)
    
    try:
        buffer = generate_audit_report_pdf(logs, workspace_id, chain_status)
        filename = f"blackmess_audit_{workspace_id}_{timezone.now().strftime('%Y%m%d')}.pdf"
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_audit_excel(request):
    """Export audit log ke Excel untuk OJK/BI"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    workspace_id = request.query_params.get('workspace_id', 'default')
    days = int(request.query_params.get('days', 30))
    
    since = timezone.now() - timedelta(days=days)
    logs = ImmutableAuditLog.objects.filter(
        workspace_id=workspace_id,
        created_at__gte=since
    ).order_by('-created_at')

    chain_status = verify_audit_chain(workspace_id)

    wb = openpyxl.Workbook()
    
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1['A1'] = "LAPORAN AUDIT BLACKMESS"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = f"Workspace: {workspace_id}"
    ws1['A3'] = f"Periode: {days} hari terakhir"
    ws1['A4'] = f"Generated: {timezone.now().strftime('%d/%m/%Y %H:%M WIB')}"
    ws1['A5'] = f"Integritas Chain: {'VALID' if chain_status['valid'] else 'RUSAK!'}"
    ws1['A5'].font = Font(color="00AA00" if chain_status['valid'] else "FF0000", bold=True)
    ws1['A6'] = f"Total Log: {logs.count()}"
    ws1['A7'] = "Standar: OJK POJK No.11/POJK.03/2022"

    # Sheet 2: Audit Logs
    ws2 = wb.create_sheet("Audit Log")
    headers = ['No', 'Pengirim', 'Penerima', 'Aksi', 'Channel', 'IP Address', 'Waktu', 'Message Hash', 'Chain Hash']
    
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    for row, log in enumerate(logs, 2):
        ws2.cell(row=row, column=1, value=row-1)
        ws2.cell(row=row, column=2, value=log.sender_id)
        ws2.cell(row=row, column=3, value=log.receiver_id)
        ws2.cell(row=row, column=4, value=log.action)
        ws2.cell(row=row, column=5, value=log.channel)
        ws2.cell(row=row, column=6, value=log.ip_address or '-')
        ws2.cell(row=row, column=7, value=str(log.created_at)[:19])
        ws2.cell(row=row, column=8, value=log.message_hash)
        ws2.cell(row=row, column=9, value=log.chain_hash)

    # Sheet 3: Chain Verification
    ws3 = wb.create_sheet("Chain Verification")
    ws3['A1'] = "VERIFIKASI INTEGRITAS AUDIT CHAIN"
    ws3['A1'].font = Font(bold=True, size=12)
    ws3['A3'] = "Status"
    ws3['B3'] = "VALID" if chain_status['valid'] else "RUSAK - KEMUNGKINAN MANIPULASI!"
    ws3['B3'].font = Font(color="00AA00" if chain_status['valid'] else "FF0000", bold=True)
    ws3['A4'] = "Total Entries"
    ws3['B4'] = chain_status.get('count', 0)
    ws3['A5'] = "Algoritma"
    ws3['B5'] = "SHA-256 Blockchain-style"
    ws3['A6'] = "Keterangan"
    ws3['B6'] = chain_status.get('message', '')

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"blackmess_audit_{workspace_id}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_audit_json(request):
    """Export audit log ke JSON terenkripsi untuk arsip"""
    workspace_id = request.query_params.get('workspace_id', 'default')
    days = int(request.query_params.get('days', 30))
    
    since = timezone.now() - timedelta(days=days)
    logs = list(ImmutableAuditLog.objects.filter(
        workspace_id=workspace_id,
        created_at__gte=since
    ).order_by('created_at').values())

    chain_status = verify_audit_chain(workspace_id)
    
    export_data = {
        'metadata': {
            'workspace_id': workspace_id,
            'exported_at': timezone.now().isoformat(),
            'period_days': days,
            'total_logs': len(logs),
            'chain_valid': chain_status['valid'],
            'standard': 'OJK POJK No.11/POJK.03/2022',
            'export_hash': hashlib.sha256(
                json.dumps(logs, default=str).encode()
            ).hexdigest()
        },
        'audit_chain_status': chain_status,
        'logs': logs
    }
    
    filename = f"blackmess_audit_{workspace_id}_{timezone.now().strftime('%Y%m%d')}.json"
    response = HttpResponse(
        json.dumps(export_data, default=str, indent=2),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
